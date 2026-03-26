"""
Document ingestion pipeline for RAG.

Handles multiple file formats: PDF, CSV, JSON, XML, TXT, DOCX, images (OCR).
Extracts text, chunks it, generates embeddings, and stores in vector DB.
"""

from __future__ import annotations

import io
import json
import logging
import uuid
from typing import Any

from tova_core.rag.chunker import chunk_text, chunk_json, chunk_csv, chunk_xml
from tova_core.rag.embedder import Embedder
from tova_core.providers.vector_store import BaseVectorStore

logger = logging.getLogger(__name__)


class IngestionPipeline:
    """End-to-end document ingestion: extract -> chunk -> embed -> store.

    Usage:
        pipeline = IngestionPipeline(vector_store, embedder)
        result = await pipeline.ingest(
            content=pdf_bytes,
            content_type="application/pdf",
            collection_name="ds_user123_dataset456",
            source_name="report.pdf",
        )
    """

    def __init__(
        self,
        vector_store: BaseVectorStore,
        embedder: Embedder,
        chunk_size: int = 512,
        chunk_overlap: int = 50,
    ):
        self.vector_store = vector_store
        self.embedder = embedder
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

    async def ingest(
        self,
        content: bytes | str,
        content_type: str,
        collection_name: str,
        source_name: str = "",
        metadata: dict[str, Any] | None = None,
    ) -> dict:
        """Ingest a document into the vector store.

        Args:
            content: File content (bytes for binary, str for text)
            content_type: MIME type or file extension
            collection_name: Target vector collection
            source_name: Original filename
            metadata: Additional metadata to attach to all chunks

        Returns:
            {"chunks_created": int, "collection": str, "source": str}
        """
        base_metadata = {**(metadata or {}), "source": source_name}

        # Step 1: Extract text
        text_content = await self._extract_text(content, content_type)

        # Step 2: Chunk
        chunks = self._chunk(text_content, content_type, base_metadata)

        if not chunks:
            return {"chunks_created": 0, "collection": collection_name, "source": source_name}

        # Step 3: Ensure collection exists
        try:
            await self.vector_store.create_collection(
                collection_name, dimension=self.embedder.dimension
            )
        except Exception:
            pass  # Collection may already exist

        # Step 4: Generate embeddings
        texts = [c["text"] for c in chunks]
        embeddings = await self.embedder.embed_texts(texts)

        # Step 5: Store in vector DB
        documents = []
        for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            documents.append({
                "id": f"{source_name}_{i}_{uuid.uuid4().hex[:8]}",
                "text": chunk["text"],
                "embedding": embedding,
                "metadata": chunk.get("metadata", {}),
            })

        count = await self.vector_store.upsert(collection_name, documents)

        logger.info(
            f"Ingested {count} chunks from '{source_name}' into '{collection_name}'"
        )

        return {
            "chunks_created": count,
            "collection": collection_name,
            "source": source_name,
        }

    async def _extract_text(self, content: bytes | str, content_type: str) -> str | Any:
        """Extract text from various file formats.

        Returns either a string (for text formats) or parsed data (for structured formats).
        """
        ct = content_type.lower()

        # Plain text
        if ct in ("text/plain", ".txt", "txt"):
            return content if isinstance(content, str) else content.decode("utf-8", errors="replace")

        # JSON
        if ct in ("application/json", ".json", "json"):
            text = content if isinstance(content, str) else content.decode("utf-8")
            return json.loads(text)

        # CSV
        if ct in ("text/csv", ".csv", "csv"):
            return content if isinstance(content, str) else content.decode("utf-8")

        # XML
        if ct in ("application/xml", "text/xml", ".xml", "xml"):
            return content if isinstance(content, str) else content.decode("utf-8")

        # PDF
        if ct in ("application/pdf", ".pdf", "pdf"):
            return self._extract_pdf(content)

        # DOCX
        if ct in (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".docx", "docx",
        ):
            return self._extract_docx(content)

        # Excel
        if ct in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xlsx", "xlsx",
        ):
            return self._extract_xlsx(content)

        # Images (OCR)
        if ct.startswith("image/") or ct in (".png", ".jpg", ".jpeg", "png", "jpg"):
            return await self._extract_image_ocr(content)

        # Fallback: try as text
        if isinstance(content, str):
            return content
        try:
            return content.decode("utf-8", errors="replace")
        except Exception:
            raise ValueError(f"Unsupported content type: {content_type}")

    def _extract_pdf(self, content: bytes) -> str:
        """Extract text from PDF using pypdf."""
        try:
            from pypdf import PdfReader
        except ImportError:
            raise ImportError("Install pypdf: pip install pypdf")

        reader = PdfReader(io.BytesIO(content if isinstance(content, bytes) else content.encode()))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages)

    def _extract_docx(self, content: bytes) -> str:
        """Extract text from DOCX using python-docx."""
        try:
            from docx import Document
        except ImportError:
            raise ImportError("Install python-docx: pip install python-docx")

        doc = Document(io.BytesIO(content if isinstance(content, bytes) else content.encode()))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def _extract_xlsx(self, content: bytes) -> str:
        """Extract data from Excel as CSV-formatted text."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Install openpyxl: pip install openpyxl")

        wb = load_workbook(io.BytesIO(content if isinstance(content, bytes) else content.encode()))
        output = io.StringIO()
        import csv
        writer = csv.writer(output)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(cell) if cell is not None else "" for cell in row])
        return output.getvalue()

    async def _extract_image_ocr(self, content: bytes) -> str:
        """Extract text from images using OCR (pytesseract)."""
        try:
            from PIL import Image
            import pytesseract
        except ImportError:
            raise ImportError(
                "Install Pillow and pytesseract: pip install Pillow pytesseract"
            )

        image = Image.open(io.BytesIO(content if isinstance(content, bytes) else content.encode()))
        text = pytesseract.image_to_string(image)
        return text

    def _chunk(
        self,
        content: str | Any,
        content_type: str,
        metadata: dict[str, Any],
    ) -> list[dict]:
        """Route to the appropriate chunking strategy."""
        ct = content_type.lower()

        # JSON: use structured chunking
        if ct in ("application/json", ".json", "json") and not isinstance(content, str):
            return chunk_json(content, metadata=metadata)

        # CSV: use row-based chunking
        if ct in ("text/csv", ".csv", "csv"):
            return chunk_csv(content, metadata=metadata)

        # XML: use element-based chunking
        if ct in ("application/xml", "text/xml", ".xml", "xml"):
            return chunk_xml(content, metadata=metadata)

        # Everything else: text chunking
        text = content if isinstance(content, str) else json.dumps(content, default=str)
        return chunk_text(
            text,
            chunk_size=self.chunk_size,
            chunk_overlap=self.chunk_overlap,
            metadata=metadata,
        )
